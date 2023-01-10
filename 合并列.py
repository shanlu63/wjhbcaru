from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter


wb = load_workbook ('C:\\Users\\s.lu\\Downloads\\123\\TABLEAUINFO.xlsx')  # 打开PYTHON创建的数据文件
ws = wb['Sheet1']                                                         # 找到表格第一页
sheet = wb.active                                                         # 编辑第一页数据
a = sheet.max_row                                                         # 第一页的最大行数
b=[]
k=[]
j=[]
e=[]
f=[]

i=0

                                                                      # 建立i的初始值
while i < a:                                                              # while 循环， 当i 小于 最大行数时
    i = i+1                                                                 # 找到表格第一行
    k.insert(i, ws.cell (i,8).value)                                      # 将值放入 e 字典
    j.insert(i, ws.cell (i,9).value) 
    
      
    e.insert(i, ws.cell (i,3).value)                                      # 将值放入 e 字典  第3列的值  SECTION
    f.insert(i, ws.cell (i,4).value)                                      #将值放入 f 字典   第4列的值  numero
    



def remove_None_value_elements(k):
    """
    去除字典里的None，不需要修改任何代码
    remove the element(key/value) from dict if the value is None
    :param input_dict:
    :return: new dict 
    """
    if type(k) is not dict:
        return None
    result = {}
    for key in k:
        tmp = {}
        if k[key] is not None:
            if type(k[key]).__name__ == 'dict':
                tmp.update({key: remove_None_value_elements(k[key])})
            else:
                tmp.update({key: k[key]})
        result.update(tmp)
    return result

print(k)



wbtest = load_workbook ('C:\\Users\\s.lu\\Downloads\\model\\MODELE_TABLEAU_INFORMATION_ETUDE_PROSPECTION.xlsx')         # 打开地产分析的表格模板 
ws = wbtest['Sheet1']                                                                                                 # 进入页面1
sheet = wb.active                                                                                                     # 编辑页面
i = 0


while i < a :

    g= str( e[i])
    m= str( f[i])
    ws.cell (i+1,9).value = g+m
    print(ws.cell (i+1,9).value)
    
    
    
    l= str( k[i])
    o= str( j[i])
    ws.cell (i+1,11).value = l+'/'+o
    print(ws.cell (i+1,11).value)
    i = i+1                                               # 循环



wbtest.save ('C:\\Users\\s.lu\\Downloads\\model\\MODELE_TABLEAU_INFORMATION_ETUDE_PROSPECTION.xlsx')                  # 保存文件



