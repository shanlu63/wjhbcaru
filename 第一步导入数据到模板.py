import os
import xlrd2
import xlsxwriter
import glob

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import pandas as pd



biao_tou = "NULL"
wei_zhi = "NULL"


# 获取要合并的所有exce表格
def get_exce():
    global wei_zhi
    wei_zhi = 'C:\\Users\\s.lu\\Downloads\\123'
    all_exce = glob.glob(wei_zhi + "\*.xlsx")
    print("该目录下有" + str(len(all_exce)) + "个excel文件：")
    if (len(all_exce) == 0):
        return 0
    else:
        for i in range(len(all_exce)):
            print(all_exce[i])
        return all_exce


# 打开Exce文件
def open_exce(name):
    fh = xlrd2.open_workbook(name)
    return fh


# 获取exce文件下的所有sheet
def get_sheet(fh):
    sheets = fh.sheets()
    return sheets


# 获取sheet下有多少行数据
def get_sheetrow_num(sheet):
    return sheet.nrows


# 获取sheet下的数据
def get_sheet_data(sheet, row):
    for i in range(row):
        if (i == 0):
            global biao_tou
            biao_tou = sheet.row_values(i)
            continue
        values = sheet.row_values(i)
        all_data1.append(values)

    return all_data1


if __name__ == '__main__':
    all_exce = get_exce()
    # 得到要合并的所有exce表格数据
    if (all_exce == 0):
        print("该目录下无.xlsx文件！请检查您输入的目录是否有误！")
        os.system('pause')
        exit()

    all_data1 = []
    # 用于保存合并的所有行的数据

    # 下面开始文件数据的获取
    for exce in all_exce:
        fh = open_exce(exce)
        # 打开文件
        sheets = get_sheet(fh)
        # 获取文件下的sheet数量

        for sheet in range(len(sheets)):
            row = get_sheetrow_num(sheets[sheet])
            # 获取一个sheet下的所有的数据的行数

            all_data2 = get_sheet_data(sheets[sheet], row)
            # 获取一个sheet下的所有行的数据

    all_data1.insert(0, biao_tou)
    # 表头写入

    # 下面开始文件数据的写入
    new_exce = wei_zhi + "\TABLEAUINFO.xlsx"
    # 新建的exce文件名字

    fh1 = xlsxwriter.Workbook(new_exce)
    # 新建一个exce表

    new_sheet = fh1.add_worksheet()
    # 新建一个sheet表

    for i in range(len(all_data1)):
        for j in range(len(all_data1[i])):
            c = all_data1[i][j]
            new_sheet.write(i, j, c)

    fh1.close()
    # 关闭该exce表

    print("文件合并成功,请查看“" + wei_zhi + "”目录下TABLEAUINFO.xlsx文件！")

  
   
wb = load_workbook ('C:\\Users\\s.lu\\Downloads\\123\\TABLEAUINFO.xlsx')  # 打开PYTHON创建的数据文件
ws = wb['Sheet1']                                                         # 找到表格第一页
sheet = wb.active                                                         # 编辑第一页数据
a = sheet.max_row                                                         # 第一页的最大行数
b=[]                                                                      # 建立地点字典数据
c=[]                                                                      # 建立面积字典数据
d=[]                                                                      # 建立ZONGE字典数据
e=[]
f=[]
k=[]
j=[]

i=0  
                                                                       # 建立i的初始值
while i < a:                                                              # while 循环， 当i 小于 最大行数时
    i = i+1                                                               # 找到表格第一行
    print(ws.cell (i,5).value)                                            # 打印表格第一行， 第五列的值 地址
    b.insert(i, ws.cell (i,5).value)                                      # 将值放入 b 字典
    print(ws.cell (i,7).value)                                            # 打印表格第一行， 第7列的值  面积
    c.insert(i, ws.cell (i,7).value)                                      # 将值放入 c 字典
    print(ws.cell (i,9).value)                                            # 打印表格第一行， 第10列的值  ZONG
   
   
    k.insert(i, ws.cell (i,8).value)                                      # 将值放入 e 字典
    j.insert(i, ws.cell (i,9).value) 
    
     
    e.insert(i, ws.cell (i,3).value)                                      # 将值放入 e 字典  第3列的值  SECTION
    f.insert(i, ws.cell (i,4).value)                                      #将值放入 f 字典   第4列的值  numero
    
def remove_None_value_elements(k):
    """
    去除字典里的None，不需要修改任何代码
    remove the element(key/value) from dict if the value is None
    :param input_dict:
    :return: new dict 
    """
    if type(k) is not dict:
        return None
    result = {}
    for key in k:
        tmp = {}
        if k[key] is not None:
            if type(k[key]).__name__ == 'dict':
                tmp.update({key: remove_None_value_elements(k[key])})
            else:
                tmp.update({key: k[key]})
        result.update(tmp)
    return result

print(k)


    
wbtest = load_workbook ('C:\\Users\\s.lu\\Downloads\\model\\MODELE_TABLEAU_INFORMATION_ETUDE_PROSPECTION.xlsx')         # 打开地产分析的表格模板 

ws = wbtest['Sheet1']                                                                                                 # 进入页面1
sheet = wb.active                                                                                                     # 编辑页面
i = 0                                                                                                                 # 建立I 的初始值
while i < a:
    ws.cell (i+1,3).value = b[i]                                                                                     # 第2行，第三列开始导入数据 地址
    print(ws.cell (i+1,3).value)                                                                                     # 检查数据，可以删除
    ws.cell (i+1,10).value = c[i]                                                                                     # 第2行，第10列开始导入数据 面积
    print(ws.cell (i+1,10).value)
    
    
    
    g= str( e[i])
    m= str( f[i])
    ws.cell (i+1,9).value = g+m
    print(ws.cell (i+1,9).value)
    
    l= str( k[i])
    o= str( j[i])
    ws.cell (i+1,11).value = l+'/'+o
    print(ws.cell (i+1,11).value)
    
    i = i+1                                                                                                          # 循环
    

wbtest.save ('C:\\Users\\s.lu\\Downloads\\model\\MODELE_TABLEAU_INFORMATION_ETUDE_PROSPECTION.xlsx')                  # 保存文件
