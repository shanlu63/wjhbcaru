from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import pandas as pd


wb = load_workbook ('C:\\Users\\s.lu\\Downloads\\123\\TABLEAUINFO.xlsx')  # 打开PYTHON创建的数据文件
ws = wb['Sheet1']                                                         # 找到表格第一页
sheet = wb.active                                                         # 编辑第一页数据
a = sheet.max_row                                                         # 第一页的最大行数
b=[]                                                                      # 建立地点字典数据
c=[]                                                                      # 建立面积字典数据
d=[]                                                                      # 建立ZONGE字典数据
e=[]
f=[]
k=[]
j=[]

i=0  
                                                                       # 建立i的初始值
while i < a:                                                              # while 循环， 当i 小于 最大行数时
    i = i+1                                                               # 找到表格第一行
    print(ws.cell (i,5).value)                                            # 打印表格第一行， 第五列的值 地址
    b.insert(i, ws.cell (i,5).value)                                      # 将值放入 b 字典
    print(ws.cell (i,7).value)                                            # 打印表格第一行， 第7列的值  面积
    c.insert(i, ws.cell (i,7).value)                                      # 将值放入 c 字典
    print(ws.cell (i,9).value)                                            # 打印表格第一行， 第10列的值  ZONG
   
   
    k.insert(i, ws.cell (i,8).value)                                      # 将值放入 e 字典
    j.insert(i, ws.cell (i,9).value) 
    
     
    e.insert(i, ws.cell (i,3).value)                                      # 将值放入 e 字典  第3列的值  SECTION
    f.insert(i, ws.cell (i,4).value)                                      #将值放入 f 字典   第4列的值  numero
    
wbtest = load_workbook ('C:\\Users\\s.lu\\Downloads\\123\\MODELE_TABLEAU_INFORMATION_ETUDE_PROSPECTION.xlsx')         # 打开地产分析的表格模板 

ws = wbtest['Sheet1']                                                                                                 # 进入页面1
sheet = wb.active                                                                                                     # 编辑页面
i = 0                                                                                                                 # 建立I 的初始值
while i < a:
    ws.cell (i+1,3).value = b[i]                                                                                     # 第2行，第三列开始导入数据 地址
    print(ws.cell (i+1,3).value)                                                                                     # 检查数据，可以删除
    ws.cell (i+1,10).value = c[i]                                                                                     # 第2行，第10列开始导入数据 面积
    print(ws.cell (i+1,10).value)
    
    
    
    g= str( e[i])
    m= str( f[i])
    ws.cell (i+1,9).value = g+m
    print(ws.cell (i+1,9).value)
    
    l= str( k[i])
    o= str( j[i])
    ws.cell (i+1,11).value = l+'/'+o
    print(ws.cell (i+1,11).value)
    
    i = i+1                                                                                                          # 循环
    

wbtest.save ('C:\\Users\\s.lu\\Downloads\\123\\MODELE_TABLEAU_INFORMATION_ETUDE_PROSPECTION.xlsx')                  # 保存文件




